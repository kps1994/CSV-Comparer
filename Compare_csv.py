import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import openpyxl.utils
from datetime import datetime

# Define the path to your CSV files
csv_folder = r'C:\Users\a247777\OneDrive - Syneos Health\Documents\Py\CSV Compare'
encoding = 'ISO-8859-1'
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
output_file_name = "Comparison_Report"
Output_file_path = r'C:\Users\a247777\OneDrive - Syneos Health\Documents\Py\CSV Compare\\'
complete_file_name = f"{Output_file_path}{output_file_name}_{timestamp}.xlsx"
# Create an Excel workbook and select the active sheet
workbook = Workbook()
sheet = workbook.active

# Write headers to the Excel sheet
sheet.append([
    'S.No',
    'Table name',
    'OLD file name',
    'NEW file name',
    'Column Count from OLD file',
    'Column Count from NEW file',
    'Column count matches',
    'Row Count from OLD file',
    'Row Count from NEW file',
    'Row count matches',
    'Row count with data mismatch',
    'Cells with mismatch'
])

# Initialize row counter
row_num = 2
print('Checking the folder for csv files...')
# Loop through the CSV files in the folder
for filename in os.listdir(csv_folder):
    if filename.startswith('OLD_'):
        # Extract the table name from the file name
        print('CSV files found...')
        table_name = filename[4:]
        print('Table picked...\n' + table_name[:-4])
        # Load CSV data from both servers
        OLD_df = pd.read_csv(os.path.join(csv_folder, filename), encoding=encoding, low_memory=False)
        NEW_df = pd.read_csv(os.path.join(csv_folder, f'NEW_{table_name}'), encoding=encoding, low_memory=False)

        # Sort the DataFrames by the primary key (first column)
        OLD_df.sort_values(by=OLD_df.columns[0], inplace=True)
        NEW_df.sort_values(by=NEW_df.columns[0], inplace=True)

        # Get column and row counts
        col_count_OLD = len(OLD_df.columns)
        col_count_NEW = len(NEW_df.columns)
        row_count_OLD = len(OLD_df)
        row_count_NEW = len(NEW_df)

        # Check if column counts match
        col_count_matches = col_count_OLD == col_count_NEW

        # Check if row counts match
        row_count_matches = row_count_OLD == row_count_NEW

        # Initialize the count of rows with data mismatch
        row_count_data_mismatch = 0
        mismatch_cells = []

        # Iterate through rows and compare data
        print('cell by cell comparison in progress...')
        for row_index, (row1, row2) in enumerate(zip(OLD_df.values, NEW_df.values), 1):
            row_flag = False
            for col_index, (value1, value2) in enumerate(zip(row1, row2), 1):
                if not pd.isna(value1) and not pd.isna(value2) and value1 != value2:
                    row_flag = True
                    cell_label = openpyxl.utils.get_column_letter(col_index) + str(row_index + 1)
                    mismatch_cells.append(cell_label)
            if row_flag:
                row_count_data_mismatch += 1
        # Join the mismatch cell labels into a comma-separated string
        mismatch_cells_str = ', '.join(mismatch_cells)


        # Write data to Excel sheet
        print('writing data to output report...')
        sheet.cell(row=row_num, column=1, value=row_num - 1)
        sheet.cell(row=row_num, column=2, value=table_name[:-4])
        sheet.cell(row=row_num, column=3, value=filename)
        sheet.cell(row=row_num, column=4, value=f'NEW_{table_name}')
        sheet.cell(row=row_num, column=5, value=col_count_OLD)
        sheet.cell(row=row_num, column=6, value=col_count_NEW)
        sheet.cell(row=row_num, column=7, value='Yes' if col_count_matches else 'No')
        sheet.cell(row=row_num, column=8, value=row_count_OLD)
        sheet.cell(row=row_num, column=9, value=row_count_NEW)
        sheet.cell(row=row_num, column=10, value='Yes' if row_count_matches else 'No')
        sheet.cell(row=row_num, column=11, value=row_count_data_mismatch)
        sheet.cell(row=row_num, column=12, value=mismatch_cells_str)

        # Increment row counter
        row_num += 1

print('Formatting output sheet...')
# Format header row with bold font
for cell in sheet[1]:
    cell.font = Font(bold=True)
# Apply conditional formatting to cells in the "Row count with data mismatch" column
for row in sheet.iter_rows(min_row=2, max_row=row_num, min_col=11, max_col=11):
    for cell in row:
        if cell.value is not None and cell.value > 0:
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
# Apply conditional formatting to 'Row count matches' and 'Column count matches' columns
for row in sheet.iter_rows(min_row=2, max_row=row_num, min_col=7, max_col=10):
    for cell in row:
        if cell.value == 'No':
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Save the Excel workbook
print('Save in progress...')
workbook.save(complete_file_name)

print('Comparison report generated successfully.')
